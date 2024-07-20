from Schematic_design.component_info import Ui_Form
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *    #声明无参数的信号
import sys
import xlrd
import csv
import pandas as pd
import openpyxl # 导入openpyxl模块


class import_ui(QWidget, Ui_Form):
    signal_1 = pyqtSignal(str, str, list)
    def __init__(self):
        super(import_ui, self).__init__()
        self.ui = Ui_Form()
        self.setupUi(self)
        self.setWindowTitle("导入")

        self.pin_info_list = None
        self.designator = None
        self.comment = None

        self.file_pushButton.clicked.connect(self.slot_grid_opendir)
        self.confirm_pushButton.clicked.connect(self.return_info)

    def slot_grid_opendir(self):
        global grid_filepath

        grid_filepath, fileType = QFileDialog.getOpenFileName(self.file_pushButton, "打开文件夹",
                                                              "../result",
                                                              "*.txt;;*.xls;;*.xlsx;;All Files(*)")
        if grid_filepath == '':
            pass  # 防止关闭或取消导入关闭所有页面
        else:
            print('filepath:', grid_filepath)
            print('filetype:', fileType)

            if fileType == '*.txt':
                self.pin_info_list = self.slot_open_txt()

            elif fileType == '*.xlsx':
                print("正在打开xlsx")
                self.pin_info_list = self.slot_open_xlsx()
            elif fileType == '*.xls':
                print("正在打开excel")
                self.pin_info_list = self.slot_open_xls()

        self.file_Edit.setText(grid_filepath)

    def slot_open_xlsx(self):
        print("open xlsx")
        wb = openpyxl.load_workbook(grid_filepath)  # 创建workbook对象
        ws = wb.active  # 得到当前活跃表单的对象（打开该xlsx文件，直接看到的表单就为活跃表单）
        i = 0
        double_list = []
        while True:
            row = 2 + i
            num = ws['A%d' % row].value
            name = ws['B%d' % row].value


            if name is None:
                break
            adm = [num, name]
            print('adm:', adm)
            double_list.append(adm)
            #print("%s %s" % (num, name))
            i += 1
        return double_list
        #for line in ws:
        #    num = line[0].value
        #    name = line[1].value
        #    print("%s %s" % (num, name))


    def slot_open_txt(self):
        print('open txt')
        print(grid_filepath)
        file = open(grid_filepath, mode='r', encoding='UTF-8')
        double_list = []
        contents = file.readlines()
        for msg in contents:
            msg = msg.strip('\n')
            adm = msg.split(',')
            print('adm:', adm)
            double_list.append(adm)
        file.close()
        #print(double_list)
        return double_list
        #data = np.genfromtxt(self.grid_filepath, dtype=[int, str])
        #print(data)

    def slot_open_xls(self):
        print('open xls')
        #data = pd.read_excel(grid_filepath, sheet_name=0, header=None)
        double_list = []
        excel = xlrd.open_workbook(grid_filepath)
        sheet = excel.sheet_by_index(0)
        print("成功打开")

        rows = sheet.nrows
        print('rows:', rows)
        #data = [[] for i in range(rows)]
        for i in range(0, rows):
            adm = []
            #print("value:", sheet.row_values(i))
            adm.append(str(int(sheet.row_values(i)[0])))
            adm.append(sheet.row_values(i)[1])
            double_list.append(adm)
            #data[i] = sheet.row_values(i)

        return double_list

    def return_info(self):
        self.designator = self.designator_Edit.text()
        self.comment = self.comment_Edit.text()
        self.signal_1.emit(self.designator, self.comment, self.pin_info_list)

        self.close()
        #self.closeEvent()

    def item_click(self, item):
        print("您选择了：", item.designator)
        self.designator_line.setText(item.designator)
        self.comment_line.setText(item.comment)



