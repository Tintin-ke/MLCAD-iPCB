import openpyxl
def save_data_to_excel(list, output):
    #output = open('data.xls', 'a+', encoding='gbk')
    #output.write('num\tname\n')

    for i in range(len(list)):
        for j in range(len(list[i])):
            output.write(str(list[i][j]))  #write函数不能写int类型的参数，所以使用str()转化
            output.write('\t')  # 相当于Tab一下，换一个单元格
        output.write('\n')  # 换行
    output.close()
    print("保存成功！")

def save_to_excel_ad(list, output):
    output.write("Object Kind")
    output.write('\t')  # 相当于Tab一下，换一个单元格
    output.write("pin designator")
    output.write('\t')  # 相当于Tab一下，换一个单元格
    output.write("Name")
    output.write('\t')  # 相当于Tab一下，换一个单元格
    output.write("X1")
    output.write('\t')  # 相当于Tab一下，换一个单元格
    output.write("Y1")
    output.write('\n')  # 换行
    for i in range(len(list)):
        output.write("pin")
        output.write('\t')  # 相当于Tab一下，换一个单元格
        output.write(str(list[i][0]))  #write函数不能写int类型的参数，所以使用str()转化
        output.write('\t')  # 相当于Tab一下，换一个单元格
        output.write(str(list[i][1]))  # write函数不能写int类型的参数，所以使用str()转化
        output.write('\t')
        if i < len(list) / 2:
            output.write('0')
        else :
            output.write('200')
        output.write('\t')
        y = 1000 - i * 100
        output.write(str(y))
        output.write('\n')  # 换行
    output.close()
    print("save complete")


def sava_data_to_xlsx(list, output):
    title = ['Num', 'Name']
    output['A1'] = title[0]
    output['B1'] = title[1]
    for i in range(len(list)):
        output.cell(i + 2, 1).value = int(list[i][0])
        output.cell(i + 2, 2).value = list[i][1]


'''
def sava_data_to_xlsx(list, output):
    title = ['Object Kind', 'pin designator', 'Name', 'X1', 'Y1']
    output['A1'] = title[0]
    output['B1'] = title[1]
    output['C1'] = title[2]
    output['D1'] = title[3]
    output['E1'] = title[4]
    for i in range(len(list)):
        output.cell(i + 2, 1).value = 'pin'
        output.cell(i + 2, 2).value = int(list[i][0])
        output.cell(i + 2, 3).value = list[i][1]
        y = 1000 - i * 100
        if i < len(list) / 2:
            output.cell(i + 2, 4).value = 0
            output.cell(i + 2, 5).value = y
        else:
            output.cell(i + 2, 4).value = 2000
            output.cell(i + 2, 5).value = y + len(list) / 2 * 100
'''



def save_data_to_txt(list, output):
    #output = open('data.txt', 'w', encoding='gbk')
    #output.write('name,gender,status,age\n')

    for row in list:
        rowtxt = '{},{}'.format(row[0], row[1])
        output.write(rowtxt)
        output.write('\n')
    output.close()
    print("保存成功！")